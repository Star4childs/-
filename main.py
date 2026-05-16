import datetime
import io
import csv
import os
import uuid
import httpx
from fastapi import FastAPI, Request, File, UploadFile, Form
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from fastapi.responses import StreamingResponse, RedirectResponse
import openpyxl
from openpyxl.utils import get_column_letter

app = FastAPI()

# Секретный ключ для сессий (замените на случайный перед деплоем)
app.add_middleware(SessionMiddleware, secret_key="a7f1b2c3d4e5f6a7b8c9d0e1f2a3b4c5d6e7f8a9b0c1d2e3f4a5b6c7d8e9f0")

# Папка для загрузок
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

templates = Jinja2Templates(directory="templates")

# Внешний API
EXTERNAL_API_BASE = "http://5.129.248.80:8000"

# Файл с пользователями
USERS_FILE = "users.xlsx"

# ------------- Вспомогательные функции -------------
def get_session_dir(session_id: str) -> str:
    path = os.path.join(UPLOAD_DIR, session_id)
    os.makedirs(path, exist_ok=True)
    return path

def log_action(session: dict, action: str, response: str):
    if "logs" not in session:
        session["logs"] = []
    session["logs"].append({
        "id": len(session["logs"]) + 1,
        "time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action,
        "response": response
    })

async def fetch_models(client: httpx.AsyncClient) -> list[str]:
    try:
        resp = await client.get(f"{EXTERNAL_API_BASE}/models")
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, dict) and "models" in data:
            return data["models"]
        if isinstance(data, list):
            return data
        return []
    except Exception as e:
        print(f"Ошибка получения моделей: {e}")
        return []

async def call_train_api(client: httpx.AsyncClient, model_name: str,
                         file_path: str, start_row: int, end_row: int) -> dict:
    with open(file_path, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        train_data = [row for i, row in enumerate(reader) if start_row - 1 <= i < end_row]
    payload = {"model": model_name, "data": train_data}
    resp = await client.post(f"{EXTERNAL_API_BASE}/train", json=payload)
    resp.raise_for_status()
    return resp.json()

def read_users():
    """Читает пользователей из Excel, возвращает список словарей"""
    if not os.path.exists(USERS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["ФИО", "Логин", "Пароль"])
        wb.save(USERS_FILE)
    users = []
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2]:
            users.append({
                "full_name": row[0],
                "login": row[1],
                "password": row[2]
            })
    wb.close()
    return users

def add_user(full_name: str, login: str, password: str):
    """Добавляет нового пользователя в Excel"""
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb["Users"]
    ws.append([full_name, login, password])
    wb.save(USERS_FILE)
    wb.close()

def check_login(login: str, password: str):
    """Проверяет логин/пароль, возвращает имя пользователя или None"""
    users = read_users()
    for u in users:
        if u["login"] == login and u["password"] == password:
            return u["full_name"]
    return None

def require_login(request: Request):
    """Если пользователь не залогинен — редирект на /auth"""
    if not request.session.get("logged_in"):
        return RedirectResponse(url="/auth")
    return None

# ------------- Маршруты (авторизация) -------------

@app.get("/")
async def home(request: Request):
    if request.session.get("logged_in"):
        return RedirectResponse(url="/data")
    return templates.TemplateResponse(request, "index.html", {"show_welcome": True})

@app.post("/start")
async def start(request: Request):
    # После приветствия отправляем на страницу входа/регистрации
    return RedirectResponse(url="/auth", status_code=303)

@app.get("/auth")
async def auth_page(request: Request):
    if request.session.get("logged_in"):
        return RedirectResponse(url="/data")
    return templates.TemplateResponse(request, "auth.html", {
        "request": request,
        "error": None,
        "success": None,   # ← добавить
        "tab": "login"
    })

@app.post("/auth/register")
async def auth_register(request: Request,
                        full_name: str = Form(...),
                        login: str = Form(...),
                        password: str = Form(...)):
    # Проверим, нет ли уже такого логина
    users = read_users()
    if any(u["login"] == login for u in users):
        return templates.TemplateResponse(request, "auth.html", {
    "request": request,
    "error": "Логин уже занят",
    "success": None,       # ← добавить
    "tab": "register"
})
    add_user(full_name, login, password)
    # После регистрации открываем вкладку "вход"
    return templates.TemplateResponse(request, "auth.html", {
        "request": request,
        "success": "Регистрация прошла успешно. Войдите.",
        "tab": "login"
    })

@app.post("/auth/login")
async def auth_login(request: Request,
                     login: str = Form(...),
                     password: str = Form(...)):
    full_name = check_login(login, password)
    if full_name:
        request.session["logged_in"] = True
        request.session["full_name"] = full_name
        return RedirectResponse(url="/data", status_code=303)
    else:
        return templates.TemplateResponse(request, "auth.html", {
    "request": request,
    "error": "Неверный логин или пароль",
    "success": None,       # ← добавить
    "tab": "login"
})

@app.post("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/", status_code=303)

# ------------- Защищённые маршруты (вкладки) -------------

@app.get("/data")
async def data_get(request: Request):
    r = require_login(request)
    if r: return r

    session = request.session
    session_id = session.get("session_id")
    context = {"active_tab": "data"}
    if session_id:
        file_path = os.path.join(get_session_dir(session_id), "data.csv")
        if os.path.exists(file_path):
            try:
                with open(file_path, 'rb') as f:
                    content = f.read()
                encodings = ['utf-8', 'windows-1251', 'cp1251', 'latin-1']
                text = None
                for enc in encodings:
                    try:
                        text = content.decode(enc)
                        break
                    except UnicodeDecodeError:
                        continue
                if text:
                    reader = csv.reader(io.StringIO(text))
                    headers = next(reader, None)
                    if headers:
                        all_rows = [row[:len(headers)] for row in reader]
                        n = session.get("n", 5)
                        rows_display = all_rows[:n]
                        context["headers"] = headers
                        context["rows"] = rows_display
                        context["n"] = n
                        context["filename"] = os.path.basename(file_path)
                        context["session_id"] = session_id
            except Exception:
                pass
    return templates.TemplateResponse(request, "index.html", context)

@app.post("/data")
async def data_post(request: Request, file: UploadFile = File(...), n: int = Form(5)):
    r = require_login(request)
    if r: return r

    if not file.filename.endswith('.csv'):
        log_action(request.session, f"Загрузка CSV-файла {file.filename}", "Ошибка: неверный формат")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "data",
            "error": "Пожалуйста, загрузите файл в формате CSV."
        })

    session = request.session
    if "session_id" not in session:
        session["session_id"] = str(uuid.uuid4())
    session_id = session["session_id"]
    session_dir = get_session_dir(session_id)

    file_path = os.path.join(session_dir, "data.csv")
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    encodings = ['utf-8', 'windows-1251', 'cp1251', 'latin-1']
    text = None
    for enc in encodings:
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        log_action(session, f"Загрузка CSV-файла {file.filename}", "Ошибка: кодировка не определена")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "data",
            "error": "Не удалось прочитать файл."
        })

    try:
        reader = csv.reader(io.StringIO(text))
        headers = next(reader, None)
        if not headers:
            log_action(session, f"Загрузка CSV-файла {file.filename}", "Ошибка: файл пуст")
            return templates.TemplateResponse(request, "index.html", {
                "active_tab": "data",
                "error": "Файл пуст или повреждён."
            })
        all_rows = [row[:len(headers)] for row in reader]
        rows_display = all_rows[:n]
    except Exception:
        log_action(session, f"Загрузка CSV-файла {file.filename}", "Ошибка: структура повреждена")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "data",
            "error": "Ошибка при чтении CSV."
        })

    session["n"] = n
    log_action(session, f"Загрузка CSV-файла {file.filename}", "ОК")
    return templates.TemplateResponse(request, "index.html", {
        "active_tab": "data",
        "headers": headers,
        "rows": rows_display,
        "n": n,
        "filename": file.filename,
        "session_id": session_id
    })

@app.get("/analytics")
async def analytics_get(request: Request):
    r = require_login(request)
    if r: return r

    session = request.session
    session_id = session.get("session_id")
    context = {"active_tab": "analytics"}

    if not session_id:
        context["no_data"] = True
        return templates.TemplateResponse(request, "index.html", context)

    file_path = os.path.join(get_session_dir(session_id), "data.csv")
    if not os.path.exists(file_path):
        context["no_data"] = True
        return templates.TemplateResponse(request, "index.html", context)

    try:
        with open(file_path, 'rb') as f:
            content = f.read()
        encodings = ['utf-8', 'windows-1251', 'cp1251', 'latin-1']
        text = None
        for enc in encodings:
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if not text:
            context["no_data"] = True
            return templates.TemplateResponse(request, "index.html", context)

        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames
        if not headers:
            context["no_data"] = True
            return templates.TemplateResponse(request, "index.html", context)

        all_rows = []
        for row in reader:
            processed_row = {}
            for k, v in row.items():
                try:
                    processed_row[k] = float(v)
                except ValueError:
                    processed_row[k] = v
            all_rows.append(processed_row)

        if not all_rows:
            context["no_data"] = True
            return templates.TemplateResponse(request, "index.html", context)

        target_col = headers[-1]
        best_row = max(all_rows, key=lambda r: r[target_col])
        worst_row = min(all_rows, key=lambda r: r[target_col])
        best_index = all_rows.index(best_row) + 1
        worst_index = all_rows.index(worst_row) + 1

        # Средние значения
        mean_row = {}
        for key in headers:
            try:
                mean_row[key] = round(sum(float(r[key]) for r in all_rows) / len(all_rows), 4)
            except (ValueError, TypeError):
                mean_row[key] = all_rows[0][key]

        context.update({
            "no_data": False,
            "headers": headers,
            "all_rows": all_rows,
            "target_col": target_col,
            "best_row": best_row,
            "best_index": best_index,
            "worst_row": worst_row,
            "worst_index": worst_index,
            "mean_row": mean_row,
        })

    except Exception:
        context["no_data"] = True

    return templates.TemplateResponse(request, "index.html", context)

@app.get("/train")
async def train_get(request: Request):
    r = require_login(request)
    if r: return r

    session = request.session
    session_id = session.get("session_id")
    models = []
    total_rows = 0
    if session_id:
        file_path = os.path.join(get_session_dir(session_id), "data.csv")
        if os.path.exists(file_path):
            with open(file_path, encoding='utf-8') as f:
                total_rows = sum(1 for _ in f) - 1
        async with httpx.AsyncClient() as client:
            models = await fetch_models(client)

    return templates.TemplateResponse(request, "index.html", {
        "active_tab": "train",
        "models": models,
        "session_id": session_id,
        "total_rows": total_rows
    })

@app.post("/train")
async def train_post(request: Request,
                     model_name: str = Form(...),
                     train_start: int = Form(...),
                     train_end: int = Form(...)):
    r = require_login(request)
    if r: return r

    session = request.session
    session_id = session.get("session_id")
    if not session_id:
        log_action(session, "Обучение", "Ошибка: нет данных")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "train",
            "error": "Нет загруженных данных."
        })

    file_path = os.path.join(get_session_dir(session_id), "data.csv")
    if not os.path.exists(file_path):
        log_action(session, "Обучение", "Ошибка: файл не найден")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "train",
            "error": "Файл данных не найден."
        })

    async with httpx.AsyncClient() as client:
        try:
            result = await call_train_api(client, model_name, file_path, train_start, train_end)
            model_id = result.get("model_id")
            session["model_id"] = model_id
            log_action(session, f"Обучение модели {model_name} (строки {train_start}-{train_end})",
                       f"ОК, model_id={model_id}")
            return templates.TemplateResponse(request, "index.html", {
                "active_tab": "train",
                "success": True,
                "model_id": model_id,
                "models": await fetch_models(client)
            })
        except Exception as e:
            log_action(session, f"Обучение модели {model_name}", f"Ошибка: {e}")
            return templates.TemplateResponse(request, "index.html", {
                "active_tab": "train",
                "error": f"Ошибка обучения: {e}",
                "models": await fetch_models(client)
            })

@app.get("/predict")
async def predict_get(request: Request):
    r = require_login(request)
    if r: return r

    session = request.session
    model_trained = "model_id" in session
    return templates.TemplateResponse(request, "index.html", {
        "active_tab": "predict",
        "model_trained": model_trained,
        "model_id": session.get("model_id", "")
    })

@app.post("/predict")
async def predict_post(request: Request):
    r = require_login(request)
    if r: return r

    session = request.session
    model_id = session.get("model_id")
    if not model_id:
        log_action(session, "Предсказание", "Ошибка: модель не обучена")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "predict",
            "error": "Модель не обучена.",
            "model_trained": False
        })

    file_path = os.path.join(get_session_dir(session["session_id"]), "data.csv")
    if not os.path.exists(file_path):
        log_action(session, "Предсказание", "Ошибка: файл данных не найден")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "predict",
            "error": "Файл данных не найден.",
            "model_trained": True
        })

    try:
        with open(file_path, encoding='utf-8') as f:
            reader = csv.DictReader(f)
            all_data = list(reader)
    except Exception as e:
        log_action(session, "Предсказание", f"Ошибка чтения CSV: {e}")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "predict",
            "error": f"Ошибка чтения CSV: {e}",
            "model_trained": True
        })

    if not all_data:
        log_action(session, "Предсказание", "Ошибка: CSV пуст")
        return templates.TemplateResponse(request, "index.html", {
            "active_tab": "predict",
            "error": "CSV-файл пуст.",
            "model_trained": True
        })

    all_columns = list(all_data[0].keys())
    target_col = all_columns[-1]
    y_true = [float(row[target_col]) for row in all_data]
    features = [{k: v for k, v in row.items() if k != target_col} for row in all_data]

    async with httpx.AsyncClient() as client:
        try:
            payload = {"model_id": model_id, "data": features}
            resp = await client.post(f"{EXTERNAL_API_BASE}/predict", json=payload)
            resp.raise_for_status()
            result = resp.json()
            y_pred = [float(val) for val in result["predictions"]]
        except Exception as e:
            log_action(session, "Предсказание", f"Ошибка API: {e}")
            return templates.TemplateResponse(request, "index.html", {
                "active_tab": "predict",
                "error": f"Ошибка предсказания: {e}",
                "model_trained": True
            })

    errors = [pred - true for true, pred in zip(y_true, y_pred)]
    mae = sum(abs(e) for e in errors) / len(errors)
    rmse = (sum(e**2 for e in errors) / len(errors)) ** 0.5
    ss_res = sum(e**2 for e in errors)
    ss_tot = sum((y - sum(y_true)/len(y_true))**2 for y in y_true)
    r2 = 1 - ss_res / ss_tot if ss_tot != 0 else 0

    log_action(session, "Предсказание", "ОК")
    return templates.TemplateResponse(request, "index.html", {
        "active_tab": "predict",
        "model_trained": True,
        "prediction_done": True,
        "y_true": y_true,
        "y_pred": y_pred,
        "errors": errors,
        "mae": round(mae, 4),
        "rmse": round(rmse, 4),
        "r2": round(r2, 4),
        "target_col": target_col,
        "model_id": model_id
    })

@app.get("/logs")
async def logs(request: Request):
    r = require_login(request)
    if r: return r

    session = request.session
    logs_list = session.get("logs", [])
    return templates.TemplateResponse(request, "index.html", {
        "active_tab": "logs",
        "logs": logs_list
    })

@app.get("/logs/data")
async def get_logs_data(request: Request):
    r = require_login(request)
    if r: return r

    session = request.session
    logs_list = session.get("logs", [])
    return {"logs": logs_list}

@app.post("/logs/clear")
async def clear_logs(request: Request):
    r = require_login(request)
    if r: return r

    request.session["logs"] = []
    return RedirectResponse(url="/logs", status_code=303)

@app.get("/logs/download")
async def download_logs(request: Request):
    r = require_login(request)
    if r: return r

    session = request.session
    logs_list = session.get("logs", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Логи действий"
    headers = ["№", "Время", "Действие пользователя", "Ответ сайта"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row_idx, log in enumerate(logs_list, 2):
        ws.cell(row=row_idx, column=1, value=log["id"])
        ws.cell(row=row_idx, column=2, value=log["time"])
        ws.cell(row=row_idx, column=3, value=log["action"])
        ws.cell(row=row_idx, column=4, value=log["response"])
    for col in range(1, 5):
        max_width = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell in row:
                if cell:
                    max_width = max(max_width, len(str(cell)))
        ws.column_dimensions[get_column_letter(col)].width = max_width + 2
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=logs.xlsx"})